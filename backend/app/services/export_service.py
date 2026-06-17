import io
from typing import List
from datetime import date

from app.schemas.report import DailyReportRow


class ExportService:
    def export_daily_excel(self, rows: List[DailyReportRow]) -> io.BytesIO:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"

        header_fill = PatternFill("solid", fgColor="0F4C81")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(border_style="thin", color="E2E8F0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["Date", "Employee", "Code", "Project", "Activity", "Hours", "Billable", "Description"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        alt_fill = PatternFill("solid", fgColor="F8FAFC")
        for row_idx, r in enumerate(rows, 2):
            values = [
                str(r.work_date), r.employee_name, r.employee_code,
                r.project_name, r.activity_name, float(r.hours_worked),
                "Yes" if r.is_billable else "No", r.description or "",
            ]
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                if fill:
                    cell.fill = fill

        col_widths = [12, 25, 12, 30, 20, 8, 10, 40]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def export_daily_pdf(self, rows: List[DailyReportRow]) -> io.BytesIO:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        styles = getSampleStyleSheet()

        elements = []
        elements.append(Paragraph("HourHive – Daily Timesheet Report", styles["Title"]))
        elements.append(Spacer(1, 12))

        data = [["Date", "Employee", "Code", "Project", "Activity", "Hours", "Billable"]]
        for r in rows:
            data.append([
                str(r.work_date), r.employee_name, r.employee_code,
                r.project_name, r.activity_name,
                str(r.hours_worked), "Yes" if r.is_billable else "No",
            ])

        primary = colors.HexColor("#0F4C81")
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), primary),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("ALIGN", (5, 0), (5, -1), "RIGHT"),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        doc.build(elements)
        buf.seek(0)
        return buf
